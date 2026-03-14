#!/usr/bin/env python3
import csv
import psycopg2
import openpyxl
import sys
import os

def import_all_sheets():
    try:
        # Connect to PostgreSQL
        conn = psycopg2.connect(
            dbname="fbm_hub_dev",
            user="ops",
            password="26@December*o4",
            host="localhost",
            port=5432
        )
        conn.autocommit = True
        cursor = conn.cursor()
        
        total_imported = 0
        total_skipped = 0
        
        # Load Excel workbook
        wb = openpyxl.load_workbook('/Users/parthsharma/Desktop/babaclick/UK-US FBM Expenses DB.xlsx', data_only=True)
        
        print(f"📚 Found {len(wb.sheetnames)} sheets")
        print(f"Processing all sheets...\n")
        
        # Process each sheet
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            print(f"📄 Sheet {sheet_idx}/{len(wb.sheetnames)}: '{sheet_name}'")
            print(f"   Max rows: {ws.max_row}")
            
            # Convert sheet to CSV in memory
            csv_data = []
            for row in ws.iter_rows(values_only=True):
                csv_data.append(row)
            
            if len(csv_data) < 3:
                print(f"   ⏭️  Skipped (too few rows)")
                continue
            
            imported = 0
            skipped = 0
            
            # Skip header rows (first 2)
            for idx, row in enumerate(csv_data[2:], start=3):
                try:
                    if not row or not row[0]:
                        continue
                    
                    # Safely handle date fields
                    def safe_date(val):
                        if not val or str(val).strip() == '-' or str(val).strip() == '':
                            return None
                        
                        val_str = str(val).strip()
                        import re
                        
                        # Format 1: 2025-01-15 or 2025-01-15 10:30:00
                        if re.match(r'^\d{4}-\d{2}-\d{2}', val_str):
                            if len(val_str) <= 30:
                                return val_str
                        
                        # Format 2: 1/15/2025 or 01/15/2025
                        if re.match(r'^\d{1,2}/\d{1,2}/\d{4}(\s|:|$)', val_str):
                            if '-' not in val_str.split()[0]:
                                return val_str
                        
                        return None
                    
                    def safe_float(val):
                        if not val or str(val).strip() == '-' or str(val).strip() == '':
                            return 0
                        try:
                            return float(val)
                        except:
                            return 0
                    
                    def safe_int(val, default=1):
                        if not val or str(val).strip() == '-' or str(val).strip() == '':
                            return default
                        try:
                            return int(float(val))
                        except:
                            return default
                    
                    # Only require order_id to be non-empty
                    order_id = str(row[2]).strip() if len(row) > 2 else None
                    if not order_id or order_id.startswith('='):
                        skipped += 1
                        continue
                    
                    sql = """
                        INSERT INTO orders 
                        (order_id, sku, qty, product_name, order_date, delivery_date, 
                         ship_by_date, buy_link, po_id, total_sell_price, status, imported_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    
                    values = (
                        order_id,
                        str(row[5]).strip() if len(row) > 5 and row[5] and not str(row[5]).startswith('=') else None,
                        safe_int(row[6]) if len(row) > 6 else 1,
                        str(row[7]).strip() if len(row) > 7 and row[7] and not str(row[7]).startswith('=') else None,
                        safe_date(row[1]) if len(row) > 1 else None,
                        safe_date(row[26]) if len(row) > 26 else None,
                        safe_date(row[10]) if len(row) > 10 else None,
                        str(row[13]).strip() if len(row) > 13 and row[13] and not str(row[13]).startswith('=') else None,
                        str(row[14]).strip() if len(row) > 14 and row[14] and not str(row[14]).startswith('=') else None,
                        safe_float(row[9]) if len(row) > 9 else 0,
                        'pending',
                        f'import@babaclick.com ({sheet_name})'
                    )
                    
                    cursor.execute(sql, values)
                    if cursor.rowcount > 0:
                        imported += 1
                    else:
                        skipped += 1
                    
                except Exception as e:
                    skipped += 1
            
            print(f"   ✓ {imported} imported | ⏭️  {skipped} skipped")
            total_imported += imported
            total_skipped += skipped
        
        print(f"\n{'='*50}")
        print(f"✅ COMPLETE - ALL SHEETS IMPORTED")
        print(f"{'='*50}")
        print(f"📊 Total imported: {total_imported:,}")
        print(f"⏭️  Total skipped: {total_skipped:,}")
        
        # Verify
        cursor.execute("SELECT COUNT(*) FROM orders;")
        db_count = cursor.fetchone()[0]
        print(f"📦 Current database total: {db_count:,} orders")
        
        cursor.close()
        conn.close()
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        sys.exit(1)

if __name__ == "__main__":
    import_all_sheets()
