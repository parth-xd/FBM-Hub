#!/usr/bin/env python3
import psycopg2
import openpyxl
import sys

def import_sheet1_all_rows():
    try:
        conn = psycopg2.connect(
            dbname="fbm_hub_dev",
            user="ops",
            password="26@December*o4",
            host="localhost",
            port=5432
        )
        conn.autocommit = True
        cursor = conn.cursor()
        
        # Clear existing data
        cursor.execute("DELETE FROM orders;")
        print("✅ Cleared existing data")
        
        # Load ONLY Sheet 1
        wb = openpyxl.load_workbook('/Users/parthsharma/Desktop/babaclick/UK-US FBM Expenses DB.xlsx', data_only=True)
        ws = wb[wb.sheetnames[0]]  # Only first sheet
        
        print(f"📄 Sheet 1: '{wb.sheetnames[0]}'")
        print(f"📊 Total rows: {ws.max_row}")
        
        imported = 0
        skipped = 0
        
        # Safely handle date fields
        def safe_date(val):
            if not val or str(val).strip() == '-' or str(val).strip() == '':
                return None
            
            val_str = str(val).strip()
            import re
            
            # Format 1: 2025-01-15 or 2025-01-15 10:30:00
            if re.match(r'^\d{4}-\d{2}-\d{2}', val_str):
                if len(val_str) <= 30:
                    return val_str
            
            # Format 2: 1/15/2025 or 01/15/2025
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}(\s|:|$)', val_str):
                if '-' not in val_str.split()[0]:
                    return val_str
            
            return None
        
        def safe_float(val):
            if not val or str(val).strip() == '-' or str(val).strip() == '':
                return 0
            try:
                return float(val)
            except:
                return 0
        
        def safe_int(val, default=1):
            if not val or str(val).strip() == '-' or str(val).strip() == '':
                return default
            try:
                return int(float(val))
            except:
                return default
        
        # Process rows starting from row 3 (skip headers in rows 1-2)
        for row_num in range(3, ws.max_row + 1):
            try:
                row = []
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    row.append(cell.value)
                
                if not row or not row[0]:
                    skipped += 1
                    continue
                
                # Get order_id from column C (index 2)
                order_id = str(row[2]).strip() if len(row) > 2 and row[2] else None
                
                # Skip if no order_id or it's a formula
                if not order_id or order_id == 'None' or order_id.startswith('='):
                    skipped += 1
                    continue
                
                sql = """
                    INSERT INTO orders 
                    (order_id, sku, qty, product_name, order_date, delivery_date, 
                     ship_by_date, buy_link, po_id, total_sell_price, status, imported_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                values = (
                    order_id,
                    str(row[5]).strip() if len(row) > 5 and row[5] and not str(row[5]).startswith('=') else None,
                    safe_int(row[6]) if len(row) > 6 else 1,
                    str(row[7]).strip() if len(row) > 7 and row[7] and not str(row[7]).startswith('=') else None,
                    safe_date(row[1]) if len(row) > 1 else None,
                    safe_date(row[26]) if len(row) > 26 else None,
                    safe_date(row[10]) if len(row) > 10 else None,
                    str(row[13]).strip() if len(row) > 13 and row[13] and not str(row[13]).startswith('=') else None,
                    str(row[14]).strip() if len(row) > 14 and row[14] and not str(row[14]).startswith('=') else None,
                    safe_float(row[9]) if len(row) > 9 else 0,
                    'pending',
                    'import@babaclick.com'
                )
                
                cursor.execute(sql, values)
                if cursor.rowcount > 0:
                    imported += 1
                else:
                    skipped += 1
                
                if imported % 500 == 0:
                    print(f"  ✓ {imported} rows...")
                
            except Exception as e:
                skipped += 1
        
        print(f"\n✅ IMPORT COMPLETE!")
        print(f"   ✓ {imported:,} orders imported")
        print(f"   ⏭️  {skipped:,} rows skipped (no order_id)")
        
        # Verify
        cursor.execute("SELECT COUNT(*) FROM orders;")
        db_count = cursor.fetchone()[0]
        print(f"   📦 Database total: {db_count:,}")
        
        cursor.close()
        conn.close()
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    import_sheet1_all_rows()
